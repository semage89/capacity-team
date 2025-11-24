"""Serwis eksportu danych do PDF i Excel"""
from datetime import datetime, date
from io import BytesIO
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from .models import ResourceAllocation, Absence, User, Project


class ExportService:
    """Serwis do eksportu danych"""
    
    @staticmethod
    def export_allocations_excel(allocations: List[ResourceAllocation], 
                                 start_date: date, 
                                 end_date: date) -> BytesIO:
        """Eksportuje alokacje do Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Alokacje zasobów"
        
        # Nagłówki
        headers = ['Użytkownik', 'Projekt', 'Rola', 'Data rozpoczęcia', 'Data zakończenia', 
                  'Alokacja %', 'Notatki']
        ws.append(headers)
        
        # Styl nagłówków
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Dane
        for allocation in allocations:
            ws.append([
                allocation.user.display_name if allocation.user else '',
                allocation.project.name if allocation.project else '',
                allocation.role or '',
                allocation.start_date.isoformat() if allocation.start_date else '',
                allocation.end_date.isoformat() if allocation.end_date else 'Bezterminowo',
                f"{allocation.allocation_percentage}%",
                allocation.notes or ''
            ])
        
        # Auto-width kolumn
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Zapisz do BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_allocations_pdf(allocations: List[ResourceAllocation],
                               start_date: date,
                               end_date: date) -> BytesIO:
        """Eksportuje alokacje do PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Tytuł
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph("Raport alokacji zasobów", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Okres
        period_text = f"Okres: {start_date.isoformat()} - {end_date.isoformat()}"
        story.append(Paragraph(period_text, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Tabela
        data = [['Użytkownik', 'Projekt', 'Rola', 'Alokacja %', 'Okres']]
        
        for allocation in allocations:
            period = f"{allocation.start_date.isoformat()}"
            if allocation.end_date:
                period += f" - {allocation.end_date.isoformat()}"
            else:
                period += " - Bezterminowo"
            
            data.append([
                allocation.user.display_name if allocation.user else '',
                allocation.project.name if allocation.project else '',
                allocation.role or '',
                f"{allocation.allocation_percentage}%",
                period
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        story.append(table)
        
        # Stopka
        story.append(Spacer(1, 0.3*inch))
        footer_text = f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(footer_text, styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        return buffer

